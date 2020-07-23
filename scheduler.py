# PIECE SCHEDULER ATTEMPT 1.0
# Author: Amy Zhang
# Date: July 23, 2020
import openpyxl
from csp import Constraint, CSP
from dancer import DancerDirectory
from typing import Dict, List, Optional
from collections import defaultdict

SEM = 'S20'
DANCER_FILE = 'Assigned_Unassigned ' + SEM + '.xlsx'
PIECES_TO_AVAILABLE_TIMES_FILE = 'Pieces_To_Times ' + SEM + '.xlsx'
MAX_REHEARSALS_PER_TIME_FILE = 'Max_Rehearsals_Per_Time ' + SEM + '.xlsx'

def GetMaxRehearsalsPerTime() -> Dict[str, int]:
    book = openpyxl.load_workbook(MAX_REHEARSALS_PER_TIME_FILE)
    sheet = book.active
    max_rehearsals_dict = {}
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            max_rehearsals_dict[row[0]] = row[1]
        else:
            break
    return max_rehearsals_dict
 
def GetRehearsalsPerTime() -> Dict[str, List[str]]:
    book = openpyxl.load_workbook(PIECES_TO_AVAILABLE_TIMES_FILE)
    sheet = book.active
    rehearsal_times_dict = {}
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:
            piece = row[0].lower()
            rehearsal_times_dict[piece] = [row[1]]
            for i in range(2, len(row)):
                if row[i] is None:
                    break
                rehearsal_times_dict[piece].append(row[i])
                i += 1
        else:
            break
    return rehearsal_times_dict

class PieceSchedulingContraint(Constraint[str, str]):
    # This conflict represents two pieces with the same dancers, which cannot be
    # scheduled at the same time
    def __init__(self, piece1: str, piece2: str) -> None:
        super().__init__([piece1, piece2])
        self.piece1: str = piece1
        self.piece2: str = piece2
  
    def satisfied(self, assignment: Dict[str, str]) -> bool:
        # If either piece is not in the assignment then it is not
        # yet possible for their rehearsals to be conflicting
        if self.piece1 not in assignment or self.piece2 not in assignment:
         return True
        # check the rehearsal assigned to piece1 is not the same as the
        # rehearsal assigned to piece2
        return assignment[self.piece1] != assignment[self.piece2]
 
class TimeSchedulingContraint(Constraint[str, str]):
    # This conflict represents the number of pieces that can be assigne to a 
    # given time, which is determined by the maxRehearsalsPerTime dict
    def __init__(self, pieces: List[str]) -> None:
        super().__init__(pieces)
        self.maxRehearsalsPerTime: Dict[str, int] = GetMaxRehearsalsPerTime()
  
    def satisfied(self, assignment: Dict[str, str]) -> bool:
        time_count_dict = defaultdict(lambda: 0)
        for _, time in assignment.items():
            time_count_dict[time] += 1
            if time_count_dict[time]>self.maxRehearsalsPerTime[time]:
                return False 
        return True

def AddDancerPiecesConstraints() -> CSP[str, str]: 
    dancer_directory: Dict[str, List[str]] = DancerDirectory(DANCER_FILE)
    all_pieces: List[str] = dancer_directory.list_pieces()
    times: Dict[str, List[str]] = GetRehearsalsPerTime()
    # setting up the csp to add constraints
    csp: CSP[str, str] = CSP(all_pieces, times)
    piece_constraints_dict: Dict[str, List[str]] = defaultdict(lambda: [])
    for dancer, pieces in dancer_directory.Dancers.items():
        for i in range(len(pieces)):
            for j in range(i+1, len(pieces)):
                if pieces[i] not in piece_constraints_dict[pieces[j]] and pieces[j] not in piece_constraints_dict[pieces[i]]: 
                    csp.add_constraint(PieceSchedulingContraint(pieces[i], pieces[j]))
                    piece_constraints_dict[pieces[i]].append(pieces[j])
    csp.add_constraint(TimeSchedulingContraint(pieces))
    return csp


def exampleTest():
    csp = AddDancerPiecesConstraints()
    print("csp set up")
    solution: Optional[Dict[str, str]] = csp.backtracking_search()
    print("csp backtracking complete")
    if solution is None:
        print("No solution found! Here is the best solution:")
        print(csp.get_best_solution())
    else:
        print(solution)


if __name__ == "__main__":
    exampleTest()
